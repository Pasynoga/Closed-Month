import os
import re
import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk, filedialog, messagebox
from datetime import datetime
from copy import copy
from openpyxl.styles import PatternFill

def transliterate_ua(text):
    table = {
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'H', 'Ґ': 'G', 'Д': 'D', 'Е': 'E', 'Є': 'Ye', 'Ж': 'Zh', 'З': 'Z',
        'И': 'Y', 'І': 'I', 'Ї': 'Yi', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M', 'Н': 'N', 'О': 'O', 'П': 'P',
        'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U', 'Ф': 'F', 'Х': 'Kh', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch',
        'Ю': 'Yu', 'Я': 'Ya', 'Ь': '', '’': '', 'Є': 'Ye', 'Ї': 'Yi', 'Й': 'Y', 'Ґ': 'G',
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'h', 'ґ': 'g', 'д': 'd', 'е': 'e', 'є': 'ie', 'ж': 'zh', 'з': 'z',
        'и': 'y', 'і': 'i', 'ї': 'i', 'й': 'i', 'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p',
        'р': 'r', 'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
        'ю': 'iu', 'я': 'ia', 'ь': '', '’': ''
    }
    return ''.join(table.get(c, c) for c in text)

def get_filename_part(name):
    forms = ['ТОВ', 'ПП', 'АТ', 'ФОП', 'ОСББ', 'ДП', 'КП', 'ЗАТ', 'ПАТ', 'ТДВ', 'ТЗОВ']
    parts = name.split()
    if parts and parts[0].upper() in forms and len(parts) > 1:
        name_part = parts[1]
    else:
        name_part = parts[0] if parts else ''
    if re.search('[а-яА-ЯіїєґІЇЄҐ]', name_part):
        name_part = transliterate_ua(name_part)
    return name_part

def extract_metadata(sheet):
    seller = str(sheet['H2'].value).strip() if sheet['H2'].value else ''
    buyer = str(sheet['H3'].value).strip() if sheet['H3'].value else ''
    direction = sheet['E5'].value
    route = direction.replace('>', '-') if direction else 'UNKNOWN'
    date_cell = pd.to_datetime(sheet['D5'].value)
    from calendar import month_name
    month_str = month_name[date_cell.month]
    year_str = str(date_cell.year)
    contract1 = sheet['L2'].value
    contract2 = sheet['L3'].value
    return seller, buyer, direction.upper(), month_str, year_str, route, contract1, contract2, date_cell

def fill_crossborder_data(template_path, output_path, meta, delivery_df, is_import, wb):
    seller, buyer, direction, month_str, year_str, route, contract1, contract2, date_cell = meta

    if is_import:
        value_block = delivery_df.iloc[38:69, 2:26]
        hour_row = delivery_df.iloc[37, 2:26]
        date_column = delivery_df.iloc[38:69, 1]
        full_data_block = delivery_df.iloc[2:33, 2:26]
        full_hour_row = delivery_df.iloc[1, 2:26]
        full_date_column = delivery_df.iloc[2:33, 1]
    else:
        value_block = delivery_df.iloc[38:69, 30:54]
        hour_row = delivery_df.iloc[37, 30:54]
        date_column = delivery_df.iloc[38:69, 1]
        full_data_block = delivery_df.iloc[2:33, 30:54]
        full_hour_row = delivery_df.iloc[1, 30:54]
        full_date_column = delivery_df.iloc[2:33, 1]

    records = []
    for i in range(len(value_block.index)):
        for j in range(len(value_block.columns)):
            value = value_block.iat[i, j]
            if isinstance(value, str) and value.strip() == "-":
                continue
            if pd.notna(value) and value != 0:
                date_val = pd.to_datetime(date_column.iloc[i]).strftime('%d.%m.%Y')
                hour_val = hour_row.iloc[j]
                cross_val = value
                upper_val = None

                for k in range(len(full_date_column)):
                    if full_date_column.iloc[k] == date_column.iloc[i]:
                        for l in range(len(full_hour_row)):
                            if full_hour_row.iloc[l] == hour_val:
                                full_val = full_data_block.iat[k, l]
                                if pd.notna(full_val) and not (isinstance(full_val, str) and full_val.strip() == "-"):
                                    upper_val = full_val
                                break
                        break

                records.append([date_val, hour_val, cross_val, '', upper_val])

    df_out = pd.DataFrame(records, columns=["Дата", "Година", "Перетин", "", "Обсяг"])

    def clean_filename_part(part):
        return re.sub(r'[<>:"/\\|?*]', '', part)

    seller_part = clean_filename_part(get_filename_part(seller))
    buyer_part = clean_filename_part(get_filename_part(buyer))
    safe_direction = clean_filename_part(route)
    prefix = "IMPORT" if is_import else "EXPORT"
    filename_default = f"{prefix}_{seller_part}_{buyer_part}_{safe_direction}_{month_str}_{year_str}.xlsx".replace(' ', '_')

    Tk().withdraw()
    full_output_path = filedialog.asksaveasfilename(
        title="Збереження файлу",
        initialdir=output_path,
        initialfile=filename_default,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not full_output_path:
        messagebox.showinfo("Відміна", "Збереження скасовано користувачем.")
        return
    filename = os.path.basename(full_output_path)

    wb_template = load_workbook(template_path)
    ws = wb_template.active

    insert_rows = len(df_out)
    if insert_rows > 1:
        ws.insert_rows(15, amount=insert_rows - 1)
        for i in range(15, 15 + insert_rows - 1):
            for col in range(1, 10):
                cell_above = ws.cell(row=14, column=col)
                new_cell = ws.cell(row=i, column=col)
                new_cell.font = copy(cell_above.font)
                new_cell.border = copy(cell_above.border)
                new_cell.fill = copy(cell_above.fill)
                new_cell.number_format = cell_above.number_format
                new_cell.protection = copy(cell_above.protection)
                new_cell.alignment = copy(cell_above.alignment)
                if cell_above.data_type == 'f' and cell_above.value:
                    new_formula = re.sub(r'14', str(i), cell_above.value)
                    new_cell.value = new_formula
                elif cell_above.value:
                    new_cell.value = cell_above.value

    start_row = 14
    prev_date = None
    fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    current_fill = fill_blue

    for idx, row in df_out.iterrows():
        ws[f"A{start_row}"] = row["Дата"]
        ws[f"B{start_row}"] = row["Година"]
        ws[f"C{start_row}"] = row["Перетин"]

        sheet_name = row["Дата"]
        hour = int(row["Година"])
        try:
            sheet_base = wb[sheet_name]
            hour_start = str(hour-1).zfill(2)
            hour_end = str(hour).zfill(2)
            hour_str = f"{hour_start}-{hour_end}"
            found_row = None
            for r in range(8, 32):
                if str(sheet_base[f"B{r}"].value).strip() == hour_str:
                    found_row = r
                    break
            if found_row:
                if is_import:
                    val_D = sheet_base[f"F{found_row}"].value
                else:
                    val_D = sheet_base[f"Y{found_row}"].value
            else:
                val_D = None
        except Exception:
            val_D = None
        ws[f"D{start_row}"] = val_D

        val_F = None
        try:
            if found_row:
                val_F = sheet_base[f"D{found_row}"].value
        except Exception:
            val_F = None
        ws[f"F{start_row}"] = val_F

        val_G = None
        try:
            if is_import:
                val_G = sheet_base["E2"].value
            else:
                val_G = sheet_base["X2"].value
        except Exception:
            val_G = None
        ws[f"G{start_row}"] = val_G

        val_H = None
        try:
            if is_import:
                val_H = sheet_base["E3"].value
            else:
                val_H = sheet_base["X3"].value
        except Exception:
            val_H = None
        ws[f"H{start_row}"] = val_H

        val = row["Обсяг"]
        if pd.isna(val) or (isinstance(val, str) and val.strip() == "-"):
            ws[f"E{start_row}"] = 0
        else:
            ws[f"E{start_row}"] = val

        if is_import:
            ws[f"I{start_row}"] = f"=(C{start_row}*D{start_row})+(E{start_row}*(F{start_row}+G{start_row}+H{start_row}))"
        else:
            ws[f"I{start_row}"] = f"=(F{start_row}-G{start_row}-H{start_row})*E{start_row}-(C{start_row}*D{start_row})"

        if row["Дата"] != prev_date:
            current_fill = fill_white if current_fill == fill_blue else fill_blue
            prev_date = row["Дата"]

        for col in range(1, 10):
            ws.cell(row=start_row, column=col).fill = current_fill

        start_row += 1

    max_row = ws.max_row
    total_row = None
    for row in range(start_row, max_row + 1):
        if ws[f"D{row}"].value and "всього" in str(ws[f"D{row}"].value).lower():
            total_row = row
            break

    bottom_data = []
    if total_row:
        for r in range(total_row, max_row + 1):
            row_data = []
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                row_data.append((cell.value, copy(cell.font), copy(cell.border), copy(cell.fill), cell.number_format, copy(cell.protection), copy(cell.alignment)))
            bottom_data.append(row_data)

        ws.delete_rows(start_row, max_row - start_row + 1)

        for row_offset, row_data in enumerate(bottom_data):
            for col_index, (val, font, border, fill, numfmt, protection, alignment) in enumerate(row_data):
                cell = ws.cell(row=start_row + row_offset, column=col_index + 1)
                cell.value = val
                cell.font = font
                cell.border = border
                cell.fill = fill
                cell.number_format = numfmt
                cell.protection = protection
                cell.alignment = alignment

    ws['B2'] = seller
    ws['B3'] = buyer
    ws['B5'] = contract1
    ws['B6'] = contract2
    ws['B9'] = month_str
    ws['D9'] = year_str
    ws['B11'] = direction

    try:
        wb_template.save(full_output_path)
        messagebox.showinfo("Успіх", f"Збережено файл: {filename}")
    except PermissionError:
        messagebox.showerror("Помилка", f"Не вдалося зберегти файл {filename}. Можливо, він відкритий у Excel.")

def create_custom_report():
    Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title="Оберіть Excel-файл з вихідними даними",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file_path:
        messagebox.showinfo("Відміна", "Файл не обрано.")
        return

    try:
        xls = pd.ExcelFile(file_path)
        wb = load_workbook(file_path, data_only=True)
        delivery_df = xls.parse('Delivery&JAO schedule', header=None)

        import_cross = delivery_df.iloc[38:69, 2:26].apply(pd.to_numeric, errors='coerce')
        export_cross = delivery_df.iloc[38:69, 30:54].apply(pd.to_numeric, errors='coerce')
        has_import = (import_cross.fillna(0).values.sum() > 0)
        has_export = (export_cross.fillna(0).values.sum() > 0)

        template_path = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), 'templates', 'excel', 'Report to Custom_FORM.xlsx'
)

        if has_import:
            sheet = wb['Deposit IMPORT']
            meta = extract_metadata(sheet)
            fill_crossborder_data(template_path, os.path.dirname(file_path), meta, delivery_df, is_import=True, wb=wb)

        if has_export:
            sheet = wb['Deposit EXPORT']
            meta = extract_metadata(sheet)
            fill_crossborder_data(template_path, os.path.dirname(file_path), meta, delivery_df, is_import=False, wb=wb)

        if not has_import and not has_export:
            messagebox.showinfo("Немає даних", "Немає даних для імпорту чи експорту.")
    except Exception as e:
        messagebox.showerror("Помилка", f"Не вдалося створити звіт: {e}")